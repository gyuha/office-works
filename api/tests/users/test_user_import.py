"""Unit tests for the Excel bulk-import pure logic (no DB / no FastAPI)."""

from __future__ import annotations

import io

import pytest
from openpyxl import Workbook, load_workbook

from domains.users.schemas import UserCreate
from domains.users.service.user_import import (
    IMPORT_HEADERS,
    build_import_template,
    parse_import_rows,
)

pytestmark = pytest.mark.unit


def _xlsx(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _valid_row(email: str = "a@b.com", employee_no: str = "") -> list[object]:
    # column order: employee_no, name, department, rank, grade, phone, email
    return [employee_no, "김철수", "개발팀", "대리", "고급", "010-1111-2222", email]


class TestBuildImportTemplate:
    def test_build_import_template_has_canonical_header_row(self) -> None:
        data = build_import_template()
        workbook = load_workbook(io.BytesIO(data))
        sheet = workbook.active
        assert sheet is not None
        header = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        assert header[: len(IMPORT_HEADERS)] == IMPORT_HEADERS
        assert header[0] == "employee_no"


class TestParseImportRows:
    def test_parse_valid_rows_returns_row_numbered_user_creates(self) -> None:
        data = _xlsx([IMPORT_HEADERS, _valid_row("a@b.com"), _valid_row("c@d.com")])
        valid, errors = parse_import_rows(data)
        assert errors == []
        assert len(valid) == 2
        assert [row for row, _ in valid] == [2, 3]
        assert all(isinstance(user, UserCreate) for _, user in valid)

    def test_parse_blank_employee_no_becomes_none(self) -> None:
        data = _xlsx([IMPORT_HEADERS, _valid_row("a@b.com", employee_no="")])
        valid, errors = parse_import_rows(data)
        assert errors == []
        assert valid[0][1].employee_no is None

    def test_parse_provided_employee_no_is_kept(self) -> None:
        data = _xlsx([IMPORT_HEADERS, _valid_row("a@b.com", employee_no="EMP-900")])
        valid, errors = parse_import_rows(data)
        assert errors == []
        assert valid[0][1].employee_no == "EMP-900"

    def test_parse_duplicate_employee_no_in_file_reports_second(self) -> None:
        data = _xlsx(
            [
                IMPORT_HEADERS,
                _valid_row("a@b.com", employee_no="EMP-900"),
                _valid_row("c@d.com", employee_no="EMP-900"),
            ]
        )
        valid, errors = parse_import_rows(data)
        assert len(valid) == 1
        assert len(errors) == 1
        assert errors[0].row == 3
        assert "사번" in errors[0].reason

    def test_parse_missing_required_field_reports_row_error(self) -> None:
        bad = ["", "", "개발팀", "대리", "고급", "010-1", "x@y.com"]  # blank name
        data = _xlsx([IMPORT_HEADERS, bad])
        valid, errors = parse_import_rows(data)
        assert valid == []
        assert len(errors) == 1
        assert errors[0].row == 2

    def test_parse_invalid_email_reports_row_error(self) -> None:
        bad = ["", "김철수", "개발팀", "대리", "고급", "010-1", "not-an-email"]
        data = _xlsx([IMPORT_HEADERS, bad])
        valid, errors = parse_import_rows(data)
        assert valid == []
        assert len(errors) == 1
        assert errors[0].row == 2

    def test_parse_blank_row_is_skipped(self) -> None:
        blank: list[object] = [None, None, None, None, None, None, None]
        data = _xlsx([IMPORT_HEADERS, _valid_row("a@b.com"), blank, _valid_row("c@d.com")])
        valid, errors = parse_import_rows(data)
        assert errors == []
        assert len(valid) == 2

    def test_parse_duplicate_email_in_file_keeps_first_reports_second(self) -> None:
        data = _xlsx([IMPORT_HEADERS, _valid_row("dup@x.com"), _valid_row("dup@x.com")])
        valid, errors = parse_import_rows(data)
        assert len(valid) == 1
        assert len(errors) == 1
        assert errors[0].row == 3

    def test_parse_wrong_header_reports_header_error(self) -> None:
        data = _xlsx([["a", "b", "c", "d", "e", "f", "g"], _valid_row()])
        valid, errors = parse_import_rows(data)
        assert valid == []
        assert len(errors) == 1
        assert errors[0].row == 1

    def test_parse_non_xlsx_bytes_raises_value_error(self) -> None:
        with pytest.raises(ValueError, match="xlsx"):
            parse_import_rows(b"this is not an excel file")
