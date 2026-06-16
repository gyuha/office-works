"""Unit tests for scripts/users_cli.py pure helpers + template handler (no DB)."""

from __future__ import annotations

import io
from pathlib import Path

import pytest
import scripts.users_cli as cli
from openpyxl import load_workbook

from domains.users.service.user_import import IMPORT_HEADERS

pytestmark = pytest.mark.unit


class TestResolveTemplatePath:
    def test_resolve_template_path_none_returns_cwd_default(self) -> None:
        assert cli.resolve_template_path(None) == Path.cwd() / cli.DEFAULT_TEMPLATE_NAME

    def test_resolve_template_path_directory_returns_default_inside_it(
        self, tmp_path: Path
    ) -> None:
        assert cli.resolve_template_path(str(tmp_path)) == tmp_path / cli.DEFAULT_TEMPLATE_NAME

    def test_resolve_template_path_xlsx_path_returns_exact_path(self, tmp_path: Path) -> None:
        target = tmp_path / "custom.xlsx"
        assert cli.resolve_template_path(str(target)) == target


class TestDecideExitCode:
    def test_decide_exit_code_parse_failed_returns_2(self) -> None:
        assert cli.decide_exit_code(0, 0, parse_failed=True) == 2

    def test_decide_exit_code_some_failed_rows_returns_1(self) -> None:
        assert cli.decide_exit_code(3, 2, parse_failed=False) == 1

    def test_decide_exit_code_all_success_returns_0(self) -> None:
        assert cli.decide_exit_code(5, 0, parse_failed=False) == 0


class TestRunTemplate:
    def test_run_template_writes_valid_xlsx_with_canonical_header(self, tmp_path: Path) -> None:
        target = tmp_path / "out.xlsx"
        assert cli.run_template(str(target)) == 0

        sheet = load_workbook(io.BytesIO(target.read_bytes())).active
        assert sheet is not None
        header = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        assert header == IMPORT_HEADERS

    def test_run_template_overwrites_existing_file(self, tmp_path: Path) -> None:
        target = tmp_path / "out.xlsx"
        target.write_bytes(b"stale")
        assert cli.run_template(str(target)) == 0

        # The stale bytes are gone — it is a real workbook with the header row.
        sheet = load_workbook(io.BytesIO(target.read_bytes())).active
        assert sheet is not None
        header = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        assert header == IMPORT_HEADERS
