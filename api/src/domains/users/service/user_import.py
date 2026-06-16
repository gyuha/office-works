"""Excel (.xlsx) bulk-import helpers for the users directory — pure functions.

No DB and no FastAPI here: parse an uploaded workbook into validated
:class:`UserCreate` rows plus per-row errors, and build a blank import template.
Kept pure so the parsing/validation logic is unit-testable without infrastructure.
"""

from __future__ import annotations

import io

from openpyxl import Workbook, load_workbook
from pydantic import ValidationError

from domains.users.schemas import UserCreate, UserImportRowError

# Canonical column order for both the template and the importer. ``employee_no``
# is optional — blank cells are server-generated; filled values must be unique.
IMPORT_HEADERS: list[str] = [
    "employee_no",
    "name",
    "department",
    "rank",
    "grade",
    "phone",
    "email",
]


def build_import_template() -> bytes:
    """Return a blank ``.xlsx`` carrying only the canonical header row, as bytes."""
    workbook = Workbook()
    sheet = workbook.active
    if sheet is None:  # pragma: no cover - a fresh Workbook always has an active sheet
        sheet = workbook.create_sheet()
    sheet.title = "users"
    sheet.append(IMPORT_HEADERS)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _cell_to_text(cell: object) -> str:
    """Normalize a worksheet cell value to a trimmed string ('' for blanks)."""
    if cell is None:
        return ""
    return str(cell).strip()


def _first_error(exc: ValidationError) -> str:
    err = exc.errors()[0]
    loc = ".".join(str(part) for part in err.get("loc", ()))
    msg = err.get("msg", "유효하지 않은 값입니다.")
    return f"{loc}: {msg}" if loc else msg


def parse_import_rows(
    file_bytes: bytes,
) -> tuple[list[tuple[int, UserCreate]], list[UserImportRowError]]:
    """Parse an uploaded ``.xlsx`` into ``(valid (excel_row, UserCreate) rows, row errors)``.

    Excel rows are 1-based with the header on row 1, so data starts at row 2. Each
    valid row carries its Excel row number so a later DB conflict can be reported
    against the right line. Fully-blank rows are skipped. Emails duplicated within
    the file keep only the first occurrence; later duplicates become row errors.
    Raises ``ValueError`` if the bytes are not a readable ``.xlsx``.
    """
    try:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("유효한 .xlsx 파일이 아닙니다.") from exc

    sheet = workbook.active
    if sheet is None:
        workbook.close()
        return [], [UserImportRowError(row=1, reason="시트를 찾을 수 없습니다.")]

    row_iter = sheet.iter_rows(values_only=True)
    header = next(row_iter, None)
    if header is None:
        workbook.close()
        return [], [UserImportRowError(row=1, reason="빈 파일입니다.")]

    width = len(IMPORT_HEADERS)
    header_norm = [_cell_to_text(c).lower() for c in list(header)[:width]]
    if header_norm != IMPORT_HEADERS:
        workbook.close()
        return [], [
            UserImportRowError(
                row=1,
                reason=f"헤더가 올바르지 않습니다. 기대 컬럼: {', '.join(IMPORT_HEADERS)}",
            )
        ]

    valid: list[tuple[int, UserCreate]] = []
    errors: list[UserImportRowError] = []
    seen_emails: set[str] = set()
    seen_employee_nos: set[str] = set()

    for excel_row, raw in enumerate(row_iter, start=2):
        cells = list(raw)[:width]
        cells += [None] * (width - len(cells))
        if all(_cell_to_text(c) == "" for c in cells):
            continue
        data = {key: _cell_to_text(cell) for key, cell in zip(IMPORT_HEADERS, cells, strict=True)}
        try:
            user = UserCreate(**data)
        except ValidationError as exc:
            errors.append(UserImportRowError(row=excel_row, reason=_first_error(exc)))
            continue
        if user.email in seen_emails:
            errors.append(
                UserImportRowError(row=excel_row, reason=f"파일 내 이메일 중복: {user.email}")
            )
            continue
        if user.employee_no is not None and user.employee_no in seen_employee_nos:
            errors.append(
                UserImportRowError(row=excel_row, reason=f"파일 내 사번 중복: {user.employee_no}")
            )
            continue
        seen_emails.add(user.email)
        if user.employee_no is not None:
            seen_employee_nos.add(user.employee_no)
        valid.append((excel_row, user))

    workbook.close()
    return valid, errors
